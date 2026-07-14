"""
Generació d'un Excel amb un full per producte.
Cada full té: Botiga | % vs preu botiga | Data1-Preu | Data1-Notes | Data2-Preu | Data2-Notes | ...

Millores respecte la versió anterior:
- Cada data ocupa dues columnes: "Preu" i "Notes" (per apunts manuals).
- Les notes NO s'esborren en actualitzar: es llegeixen del fitxer existent
  i es tornen a escriure després de regenerar les dades.
- Els preus introduïts manualment es marquen amb to taronja i cursiva.
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import database as db

EXCEL_PATH = os.environ.get("PREUS_EXCEL_PATH", "seguiment_preus.xlsx")

MANUAL_FILL      = PatternFill(start_color="FFE8B8", end_color="FFE8B8", fill_type="solid")
NOTES_FILL       = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
DESTACADA_FILL   = PatternFill(start_color="CCFF66", end_color="CCFF66", fill_type="solid")  # verd fosforito
MANUAL_FONT   = Font(italic=True)
BOLD_FONT     = Font(bold=True)
HEADER_FONT   = Font(bold=True, color="FFFFFF")
HEADER_FILL   = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SUBHEAD_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PCT_POS_FONT  = Font(color="007A33", bold=True)
PCT_NEG_FONT  = Font(color="C00000", bold=True)
WRAP          = Alignment(wrap_text=True, vertical="top")
THIN_BORDER   = Border(
    right=Side(style="thin", color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)


def _nom_full_valid(nom):
    invalid = r'[]:*?/\\'
    for ch in invalid:
        nom = nom.replace(ch, "")
    return nom[:31] if nom else "Producte"


def _ultim_valor_no_nul(fila, columnes_data):
    valor = None
    for col in columnes_data:
        v = fila[col]
        if pd.notna(v):
            valor = v
    return valor


def _llegir_notes_existents(path, nom_full, columnes_data, botigues):
    """
    Llegeix les notes que l'usuari ha escrit a les columnes 'Notes' del fitxer
    existent ABANS de sobreescriure'l, per poder-les restaurar després.

    Retorna un dict: { (nom_botiga, data): text_nota }
    """
    notes = {}
    if not os.path.exists(path):
        return notes
    try:
        wb_old = load_workbook(path, data_only=True)
        if nom_full not in wb_old.sheetnames:
            return notes
        ws_old = wb_old[nom_full]

        # Fila 1: capçaleres principals (dates, fusionades per parelles)
        # Fila 2: sub-capçaleres "Preu" / "Notes"
        # Fila 3+: dades

        # Construïm el mapa columna → (tipus, data)
        # tipus = "preu" | "notes"
        mapa_col = {}  # col_idx (1-based) -> ("preu"/"notes", "YYYY-MM-DD")
        data_activa = None
        for col_idx in range(1, ws_old.max_column + 1):
            cap1 = ws_old.cell(row=1, column=col_idx).value
            cap2 = ws_old.cell(row=2, column=col_idx).value
            if cap1 and str(cap1).strip() not in ("Botiga", "% vs preu botiga"):
                data_activa = str(cap1).strip()
            if cap2 and data_activa:
                if str(cap2).strip() == "Notes":
                    mapa_col[col_idx] = ("notes", data_activa)

        # Construïm el mapa fila → nom_botiga
        mapa_fila = {}  # row_idx (1-based) -> nom_botiga
        for row_idx in range(3, ws_old.max_row + 1):
            val = ws_old.cell(row=row_idx, column=1).value
            if val:
                nom = str(val).strip()
                if nom.startswith("⭐"):
                    nom = nom.lstrip("⭐").strip()
                mapa_fila[row_idx] = nom

        # Recollim les notes
        for col_idx, (tipus, data) in mapa_col.items():
            if data not in columnes_data:
                continue  # data nova, no cal restaurar
            for row_idx, botiga in mapa_fila.items():
                val = ws_old.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip():
                    notes[(botiga, data)] = str(val).strip()
    except Exception:
        pass  # si el fitxer antic esta corrupte o no es pot llegir, continuem sense notes
    return notes


def exportar_excel(path=None):
    if path is None:
        path = EXCEL_PATH

    productes = db.llistar_productes()

    # Importem openpyxl directament per tenir control total sobre el fitxer
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # treure el full buit per defecte

    if not productes:
        ws = wb.create_sheet("Info")
        ws["A1"] = "Encara no hi ha productes carregats"
        wb.save(path)
        return path

    for prod in productes:
        registres = db.historic_per_producte(prod["id"])
        preu_referencia = prod["preu_referencia"]
        nom_full = _nom_full_valid(prod["nom"])

        # Obtenim quines URLs són destacades per a aquest producte
        urls_prod = db.llistar_urls(prod["id"])
        botigues_destacades = {u["botiga"] for u in urls_prod if u["destacada"]}

        if not registres:
            df_pivot  = pd.DataFrame({"Botiga": []})
            df_manual = pd.DataFrame()
            columnes_data = []
        else:
            df = pd.DataFrame(
                [(r["data_hora"], r["botiga"], r["preu"], bool(r["es_manual"]))
                 for r in registres],
                columns=["Data", "Botiga", "Preu", "Manual"]
            )
            df_pivot = df.pivot_table(
                index="Botiga", columns="Data", values="Preu", aggfunc="first"
            ).reset_index()
            df_manual = df.pivot_table(
                index="Botiga", columns="Data", values="Manual", aggfunc="first"
            ).reset_index()
            columnes_data = [c for c in df_pivot.columns if c != "Botiga"]

            if preu_referencia:
                percentatges = []
                for _, fila in df_pivot.iterrows():
                    ultim = _ultim_valor_no_nul(fila, columnes_data)
                    percentatges.append(
                        (ultim - preu_referencia) / preu_referencia if ultim is not None else None
                    )
            else:
                percentatges = [None] * len(df_pivot)

            df_pivot.insert(1, "% vs preu botiga", percentatges)
            # Les destacades sempre al principi, després per % ascendent
            df_pivot["_destacada"] = df_pivot["Botiga"].isin(botigues_destacades).astype(int)
            df_pivot = df_pivot.sort_values(
                ["_destacada", "% vs preu botiga"],
                ascending=[False, True],
                na_position="last"
            ).drop(columns=["_destacada"]).reset_index(drop=True)

        botigues = df_pivot["Botiga"].tolist() if not df_pivot.empty else []

        # Llegim les notes del fitxer anterior ABANS de sobreescriure
        notes_existents = _llegir_notes_existents(path, nom_full, columnes_data, botigues)

        ws = wb.create_sheet(nom_full)

        if df_pivot.empty:
            ws["A1"] = "Sense dades"
            continue

        n_rows = len(df_pivot)
        n_dates = len(columnes_data)

        # ── Capçalera fila 1 ──────────────────────────────────────────────
        # Col 1: Botiga  |  Col 2: %  |  Col 3+4: Data1  |  Col 5+6: Data2 ...
        for c_idx, titol in enumerate(["Botiga", "% vs preu botiga"], start=1):
            cel = ws.cell(row=1, column=c_idx, value=titol)
            cel.font = HEADER_FONT
            cel.fill = HEADER_FILL
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Fusiona les dues files de capçalera per a Botiga i %
            ws.merge_cells(start_row=1, start_column=c_idx, end_row=2, end_column=c_idx)

        for d_idx, data in enumerate(columnes_data):
            col_preu  = 3 + d_idx * 2
            col_notes = col_preu + 1

            # Fila 1: data fusionada per parella de columnes
            cel_data = ws.cell(row=1, column=col_preu, value=data)
            cel_data.font = HEADER_FONT
            cel_data.fill = HEADER_FILL
            cel_data.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=1, start_column=col_preu, end_row=1, end_column=col_notes)

            # Fila 2: sub-capçaleres Preu / Notes
            for col, text in ((col_preu, "Preu"), (col_notes, "Notes")):
                cel = ws.cell(row=2, column=col, value=text)
                cel.font = BOLD_FONT
                cel.fill = SUBHEAD_FILL
                cel.alignment = Alignment(horizontal="center")

        # ── Dades fila 3+ ────────────────────────────────────────────────
        df_manual_idx = df_manual.set_index("Botiga") if not df_manual.empty else None

        for r_idx in range(n_rows):
            fila_excel = r_idx + 3
            nom_botiga = df_pivot.iloc[r_idx]["Botiga"]

            # Col A: Botiga (fons verd fosforito si és destacada)
            es_dest = nom_botiga in botigues_destacades
            cel_bot = ws.cell(row=fila_excel, column=1, value=nom_botiga)
            cel_bot.font = BOLD_FONT
            if es_dest:
                cel_bot.fill = DESTACADA_FILL

            # Col B: % vs preu botiga
            cel_pct = ws.cell(row=fila_excel, column=2)
            if es_dest:
                cel_pct.fill = DESTACADA_FILL
            valor_pct = df_pivot.iloc[r_idx]["% vs preu botiga"]
            if pd.notna(valor_pct):
                cel_pct.value = valor_pct
                cel_pct.number_format = "+0.0%;-0.0%"
                cel_pct.font = PCT_POS_FONT if valor_pct > 0 else PCT_NEG_FONT
            else:
                cel_pct.value = "—"

            # Columnes de dates
            for d_idx, col_name in enumerate(columnes_data):
                col_preu  = 3 + d_idx * 2
                col_notes = col_preu + 1

                # Cel·la de preu
                cel_preu = ws.cell(row=fila_excel, column=col_preu)
                valor_preu = df_pivot.iloc[r_idx].get(col_name)
                if pd.notna(valor_preu):
                    cel_preu.value = float(valor_preu)
                    cel_preu.number_format = "0.00"

                    es_manual = False
                    if df_manual_idx is not None and nom_botiga in df_manual_idx.index:
                        try:
                            es_manual = bool(df_manual_idx.loc[nom_botiga, col_name])
                        except KeyError:
                            es_manual = False
                    if es_manual:
                        cel_preu.fill = MANUAL_FILL
                        cel_preu.font = MANUAL_FONT

                # Cel·la de notes: fons blau clar + restaurem el text si n'hi havia
                cel_notes = ws.cell(row=fila_excel, column=col_notes)
                cel_notes.fill = NOTES_FILL
                cel_notes.alignment = WRAP
                nota = notes_existents.get((nom_botiga, col_name))
                if nota:
                    cel_notes.value = nota

        # ── Freeze, amplades i filtre ────────────────────────────────────
        ws.freeze_panes = "C3"  # fixes: Botiga + % + fila de capçaleres
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{n_rows + 2}"

        # Amplades fixes per tipus de columna
        ws.column_dimensions["A"].width = 28   # Botiga
        ws.column_dimensions["B"].width = 16   # %
        for d_idx in range(n_dates):
            ws.column_dimensions[get_column_letter(3 + d_idx * 2)].width = 10    # Preu
            ws.column_dimensions[get_column_letter(4 + d_idx * 2)].width = 15    # Notes

        # Alçada de les files de dades
        for fila_excel in range(3, n_rows + 3):
            ws.row_dimensions[fila_excel].height = 22

    wb.save(path)
    return path


def escriure_nota_excel(nom_botiga, data, nota, path=None):
    """
    Escriu una nota a la cel·la Notes del dia indicat per a una botiga concreta,
    sense tocar res més del fitxer. Si el fitxer no existeix o no troba la
    combinació botiga+data, no fa res.

    Paràmetres:
        nom_botiga: nom de la botiga tal com apareix a la columna A de l'Excel
        data:       data en format "YYYY-MM-DD"
        nota:       text a escriure (ex: "Campanya")
        path:       ruta de l'Excel (per defecte EXCEL_PATH)
    """
    if path is None:
        path = EXCEL_PATH
    if not os.path.exists(path):
        return

    try:
        wb = load_workbook(path)
    except Exception:
        return

    for ws in wb.worksheets:
        # Busquem la columna Notes que correspon a aquesta data
        col_notes = None
        for col_idx in range(1, ws.max_column + 1):
            cap1 = ws.cell(row=1, column=col_idx).value
            cap2 = ws.cell(row=2, column=col_idx).value
            if cap1 and str(cap1).strip() == data and cap2 and str(cap2).strip() == "Notes":
                col_notes = col_idx
                break
            # Les cel·les fusionades deixen la data a la primera cel·la;
            # la sub-capçalera Notes és sempre la columna de la dreta del Preu.
            if cap2 and str(cap2).strip() == "Notes":
                # Comprova si la capçalera de data de la columna esquerra coincideix
                cap1_esq = ws.cell(row=1, column=col_idx - 1).value if col_idx > 1 else None
                if cap1_esq and str(cap1_esq).strip() == data:
                    col_notes = col_idx
                    break

        if col_notes is None:
            continue  # aquesta data no existeix en aquest full

        # Busquem la fila de la botiga
        for row_idx in range(3, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val and str(val).strip() == nom_botiga:
                cel = ws.cell(row=row_idx, column=col_notes)
                cel.value = nota
                cel.fill = NOTES_FILL
                cel.alignment = WRAP
                break

    try:
        wb.save(path)
    except PermissionError:
        pass  # si l'Excel esta obert, la nota es perdrà; l'app ja ho gestiona