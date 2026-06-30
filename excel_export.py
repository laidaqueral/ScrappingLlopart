"""
Generació d'un Excel amb un full per producte.
Cada full té: Data | Botiga1 | Botiga2 | ... amb l'evolució de preus,
i un gràfic de línies natiu d'Excel.
Els preus introduïts manualment (no extrets per scraping) es marquen
amb un to taronja i un asterisc al peu de pàgina.
"""
import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

import database as db

# Nom de fitxer FIX: sempre s'actualitza el mateix Excel, no se'n crea un nou cada vegada.
EXCEL_PATH = "seguiment_preus.xlsx"

MANUAL_FILL = PatternFill(start_color="FFE8B8", end_color="FFE8B8", fill_type="solid")
MANUAL_FONT = Font(italic=True)


def _nom_full_valid(nom):
    """Excel no permet certs caràcters ni més de 31 caràcters al nom del full."""
    invalid = r'[]:*?/\\'
    for ch in invalid:
        nom = nom.replace(ch, "")
    return nom[:31] if nom else "Producte"


def exportar_excel(path=None):
    """Genera/actualitza l'Excel. Sempre fa servir el mateix fitxer (EXCEL_PATH)
    tret que s'indiqui un altre path explícitament, de manera que cada cop
    que es crida s'actualitza el mateix document en lloc de crear-ne un nou."""
    if path is None:
        path = EXCEL_PATH
    productes = db.llistar_productes()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if not productes:
            pd.DataFrame({"Avís": ["Encara no hi ha productes carregats"]}).to_excel(
                writer, sheet_name="Info", index=False
            )

        for prod in productes:
            registres = db.historic_per_producte(prod["id"])

            if not registres:
                df_pivot = pd.DataFrame({"Botiga": []})
                df_manual = pd.DataFrame()
            else:
                df = pd.DataFrame(
                    [(r["data_hora"], r["botiga"], r["preu"], bool(r["es_manual"]))
                     for r in registres],
                    columns=["Data", "Botiga", "Preu", "Manual"]
                )
                # Files = Botiga, Columnes = Data
                df_pivot = df.pivot_table(
                    index="Botiga", columns="Data", values="Preu", aggfunc="first"
                ).reset_index()
                # Mateixa estructura però amb el flag de "manual" per saber
                # quines cel·les cal marcar visualment.
                df_manual = df.pivot_table(
                    index="Botiga", columns="Data", values="Manual", aggfunc="first"
                ).reset_index()

            nom_full = _nom_full_valid(prod["nom"])
            df_pivot.to_excel(writer, sheet_name=nom_full, index=False)

            ws = writer.sheets[nom_full]

            # Pinta de taronja i en cursiva les cel·les amb preu introduït a mà.
            if not df_manual.empty:
                cols_data = list(df_pivot.columns)
                for r_idx in range(len(df_pivot)):
                    for c_idx, col_name in enumerate(cols_data):
                        if col_name == "Botiga":
                            continue
                        try:
                            es_manual = bool(df_manual.iloc[r_idx][col_name])
                        except KeyError:
                            es_manual = False
                        if es_manual:
                            cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                            cell.fill = MANUAL_FILL
                            cell.font = MANUAL_FONT

            if not df_pivot.empty:
                ws.cell(
                    row=len(df_pivot) + 3, column=1,
                    value="🟧 Cel·les en taronja/cursiva = preu introduït manualment"
                )

            # Afegir gràfic de línies natiu si hi ha dades.
            # Ara cada FILA (botiga) és una sèrie, i les columnes (dates) són
            # les categories de l'eix X.
            if not df_pivot.empty and df_pivot.shape[1] > 1:
                n_rows = df_pivot.shape[0]
                n_cols = df_pivot.shape[1]

                chart = LineChart()
                chart.title = f"Evolució de preu — {prod['nom']}"
                chart.y_axis.title = "Preu (€)"
                chart.x_axis.title = "Data"
                chart.style = 2
                chart.width = 24
                chart.height = 12

                # Dades per files: from_rows=True, incloent la columna 1 (noms
                # de botiga) com a títol de cada sèrie.
                data = Reference(
                    ws, min_col=1, max_col=n_cols, min_row=1, max_row=n_rows + 1
                )
                chart.add_data(data, titles_from_data=True, from_rows=True)

                # Categories = capçalera de dates (fila 1, des de la columna 2)
                cats = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=1)
                chart.set_categories(cats)

                col_grafic = get_column_letter(n_cols + 2)
                ws.add_chart(chart, f"{col_grafic}2")

    return path
